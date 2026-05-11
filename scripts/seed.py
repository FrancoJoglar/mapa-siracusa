#!/usr/bin/env python3
"""
Script para cargar datos de equipos y sectores desde el consolidado de riego,
y los cuarteles desde el Excel del KMZ, en Supabase.
"""

import os
import sys
import re
from openpyxl import load_workbook
from supabase import create_client, Client

SUPABASE_URL = os.environ.get(
    "SUPABASE_URL", "https://nnelrvctqjbwfucccxfh.supabase.co"
)
SUPABASE_SERVICE_KEY = os.environ.get(
    "SUPABASE_SERVICE_KEY",
    "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6Im5uZWxydmN0cWpid2Z1Y2NjeGZoIiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc3ODI1OTgwMCwiZXhwIjoyMDkzODM1ODAwfQ.hahDH7gQo4olxi-YT1dufHsXGj1ghhEVaBlAzXfzco0",
)

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_DIR = os.path.dirname(SCRIPT_DIR)
PARENT_DIR = os.path.dirname(PROJECT_DIR)

RIEGO_PATH = os.path.join(PARENT_DIR, "Consolidado Riego 2026 -2027 .xlsx")
KMZ_EXCEL_PATH = os.path.join(PARENT_DIR, "Siracusa_2025.xlsx")

supabase: Client = create_client(SUPABASE_URL, SUPABASE_SERVICE_KEY)


def safe_str(val):
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def safe_num(val):
    if val is None:
        return None
    try:
        return float(str(val).replace(",", "."))
    except (ValueError, TypeError):
        return None


def safe_int(val):
    if val is None:
        return None
    try:
        return int(float(str(val).replace(",", ".")))
    except (ValueError, TypeError):
        return None


def parse_wkt_to_ewkt(wkt: str) -> str | None:
    if not wkt or not isinstance(wkt, str) or not wkt.startswith("POLYGON(("):
        return None
    return f"SRID=4326;{wkt}"


def main():
    # =====================================================================
    # PASO 1: Eliminar datos existentes (orden inverso por FKs)
    # =====================================================================
    print("=== Limpiando datos existentes ===")
    supabase.table("cuartel_sector").delete().neq(
        "cuartel_id", "00000000-0000-0000-0000-000000000000"
    ).execute()
    supabase.table("cuarteles").delete().neq(
        "id", "00000000-0000-0000-0000-000000000000"
    ).execute()
    supabase.table("edificaciones").delete().neq(
        "id", "00000000-0000-0000-0000-000000000000"
    ).execute()
    supabase.table("sectores").delete().neq(
        "id", "00000000-0000-0000-0000-000000000000"
    ).execute()
    supabase.table("equipos").delete().neq(
        "id", "00000000-0000-0000-0000-000000000000"
    ).execute()
    print("  Limpio.")

    # =====================================================================
    # PASO 2: Cargar equipos y sectores desde el consolidado de riego
    # =====================================================================
    print("\n=== Cargando Consolidado de Riego ===")
    wb_riego = load_workbook(RIEGO_PATH)
    ws_riego = wb_riego["E Riego"]

    equipos_creados = {}  # codigo -> id
    sector_counter = 0
    equipos_procesados = set()

    for row in ws_riego.iter_rows(min_row=2, values_only=True):
        eq_codigo = safe_int(row[0])
        sec_numero = safe_int(row[1])
        if eq_codigo is None or sec_numero is None:
            continue

        # Crear equipo si no existe
        if eq_codigo not in equipos_creados:
            nombre = f"Equipo {eq_codigo}"
            resp = (
                supabase.table("equipos")
                .upsert({"codigo": eq_codigo, "nombre": nombre}, on_conflict="codigo")
                .execute()
            )
            if resp.data:
                equipos_creados[eq_codigo] = resp.data[0]["id"]
            equipos_procesados.add(eq_codigo)

        equipo_id = equipos_creados.get(eq_codigo)
        if not equipo_id:
            print(f"  ERROR: No se pudo crear/obtener equipo {eq_codigo}")
            continue

        codigo = f"E{eq_codigo}S{sec_numero}"

        sector_data = {
            "codigo": codigo,
            "equipo_id": equipo_id,
            "numero": sec_numero,
            "caudal_nominal": safe_num(row[2]),
            "hectareas": safe_num(row[3]),
            "variedad": safe_str(row[4]),
            "caseta": safe_str(row[5]),
            "bomba": safe_str(row[6]),
            "filtro": safe_str(row[7]),
            "anio": safe_int(row[8]),
            "jefe_campo": safe_str(row[9]),
            "especie": safe_str(row[10]),
            "precipitacion": safe_num(row[11]),
            "eficiencia": safe_num(row[12]),
            "dist_entre_hilera": safe_num(row[13]),
            "dist_entre_plantas": safe_num(row[14]),
            "dist_entre_goteros": safe_num(row[15]),
            "num_lineas": safe_int(row[16]),
            "caudal_emisor": safe_num(row[17]),
            "m3_ha": safe_num(row[19]) if row[19] else None,
        }

        # Limpiar None values para upsert
        sector_data = {k: v for k, v in sector_data.items() if v is not None}

        resp = (
            supabase.table("sectores")
            .upsert(sector_data, on_conflict="codigo")
            .execute()
        )
        if resp.data:
            sector_counter += 1

    print(f"  Equipos creados: {len(equipos_creados)}")
    print(f"  Sectores creados: {sector_counter}")

    # =====================================================================
    # PASO 3: Cargar cuarteles desde el Excel del KMZ
    # =====================================================================
    print("\n=== Cargando Cuarteles desde KMZ Excel ===")
    wb_kmz = load_workbook(KMZ_EXCEL_PATH)
    ws_cuarteles = wb_kmz["Cuarteles"]

    cuarteles_count = 0
    cuarteles_errores = 0
    cuartel_ids = {}  # nombre -> id

    for row in ws_cuarteles.iter_rows(min_row=2, values_only=True):
        nombre = safe_str(row[0])
        if not nombre or not nombre.startswith("C "):
            continue

        especie = safe_str(row[1]) or ""
        variedad = safe_str(row[2]) or ""
        anio = safe_int(row[3])
        superficie = safe_num(row[4])
        plantas = safe_int(row[5])
        polinizante = safe_str(row[6]) or ""
        jefe = safe_str(row[7]) or ""
        equipo_raw = safe_str(row[8]) or ""
        sector_raw = safe_str(row[9]) or ""
        wkt = row[10] if row[10] else ""

        data = {
            "nombre": nombre,
            "especie": especie,
            "variedad": variedad if variedad else None,
            "anio_plantacion": anio,
            "superficie_ha": superficie,
            "plantas": plantas if plantas else None,
            "polinizante": polinizante if polinizante else None,
            "jefe_campo": jefe if jefe else None,
            "centro_costo": None,
            "equipo_riego": equipo_raw if equipo_raw else None,
            "sector_raw": sector_raw if sector_raw else None,
        }

        ewkt = parse_wkt_to_ewkt(wkt) if wkt else None
        if ewkt:
            data["geometria"] = ewkt

        try:
            resp = (
                supabase.table("cuarteles").upsert(data, on_conflict="nombre").execute()
            )
            if resp.data:
                cuartel_ids[nombre] = resp.data[0]["id"]
                cuarteles_count += 1
        except Exception as e:
            print(f"  ERROR en {nombre}: {e}")
            cuarteles_errores += 1

    print(f"  Cuarteles: {cuarteles_count} OK, {cuarteles_errores} errores")

    # =====================================================================
    # PASO 4: Relacionar cuarteles con sectores
    # =====================================================================
    print("\n=== Relacionando cuartel <-> sector ===")

    # Obtener sectores de BD para mapear (equipo, numero) -> sector_id
    resp_s = (
        supabase.table("sectores").select("id, codigo, equipo_id, numero").execute()
    )
    # Armar lookup: para cada sector, obtener (equipo_codigo, numero)
    sector_lookup = {}  # (equipo_codigo, numero) -> sector_id
    equipos_codigo = {}  # equipo_id -> codigo
    resp_e = supabase.table("equipos").select("id, codigo").execute()
    for e in resp_e.data or []:
        equipos_codigo[e["id"]] = e["codigo"]

    for s in resp_s.data or []:
        eq_cod = equipos_codigo.get(s["equipo_id"])
        if eq_cod and s["numero"]:
            sector_lookup[(eq_cod, s["numero"])] = s["id"]

    # Ahora para cada cuartel, parsear equipo y sector para vincular
    relaciones = 0
    for row in ws_cuarteles.iter_rows(min_row=2, values_only=True):
        nombre = safe_str(row[0])
        if not nombre or not nombre.startswith("C "):
            continue

        cuartel_id = cuartel_ids.get(nombre)
        if not cuartel_id:
            continue

        equipo_raw = safe_str(row[8]) or ""
        sector_raw = safe_str(row[9]) or ""

        # Parsear equipos y sectores como pares paralelos
        eq_nums = [int(x.strip()) for x in equipo_raw.split("-") if x.strip().isdigit()]
        sec_nums = [
            int(x.strip()) for x in sector_raw.split("-") if x.strip().isdigit()
        ]

        # Emparejar en orden
        pairs = []
        if len(eq_nums) == 1 and len(sec_nums) >= 1:
            for sn in sec_nums:
                pairs.append((eq_nums[0], sn))
        elif len(eq_nums) == len(sec_nums):
            pairs = list(zip(eq_nums, sec_nums))

        for eq_cod, sec_num in pairs:
            sector_id = sector_lookup.get((eq_cod, sec_num))
            if sector_id:
                try:
                    supabase.table("cuartel_sector").upsert(
                        {"cuartel_id": cuartel_id, "sector_id": sector_id},
                        on_conflict="cuartel_id,sector_id",
                    ).execute()
                    relaciones += 1
                except Exception as e:
                    pass

    print(f"  Relaciones creadas: {relaciones}")

    # =====================================================================
    # PASO 5: Cargar edificaciones
    # =====================================================================
    print("\n=== Cargando Edificaciones ===")
    ws_edif = wb_kmz["Edificaciones"]
    edif_count = 0

    for row in ws_edif.iter_rows(min_row=2, values_only=True):
        nombre = safe_str(row[0])
        if not nombre:
            continue
        eq = safe_str(row[1])
        sec = safe_str(row[2])
        wkt = row[3] if row[3] else ""

        data = {"nombre": nombre}
        if eq:
            data["equipo_riego"] = eq
        if sec:
            data["sector_riego"] = sec

        ewkt = parse_wkt_to_ewkt(wkt) if wkt else None
        if ewkt:
            data["geometria"] = ewkt

        supabase.table("edificaciones").insert(data).execute()
        edif_count += 1

    print(f"  Edificaciones: {edif_count}")

    print("\n✓ Seed completado con datos del consolidado de riego.")
    print(
        f"  {len(equipos_creados)} equipos | {sector_counter} sectores | {cuarteles_count} cuarteles | {relaciones} relaciones"
    )


if __name__ == "__main__":
    main()
